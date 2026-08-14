import sys
from typing import BinaryIO, Any
from ._xlsx_classify import (
    MODE_TABLE,
    MIN_TABLE_WIDTH,
    _cell_text,
    _render_caption,
    _render_out_of_span,
    _row_content_in_span,
    analyze_rows,
    classify_sheet,
    compute_sheet_stats,
    dense_4x4_rows,
    detect_region_layout,
    group_table_regions,
    is_table_row,
    render_document_row,
    render_document_sheet,
    sanitize_filename,
)
from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MissingDependencyException, MISSING_DEPENDENCY_MESSAGE
from .._stream_info import StreamInfo

# Try loading optional (but in this case, required) dependencies
# Save reporting of any exceptions for later
_xlsx_dependency_exc_info = None
try:
    import openpyxl  # noqa: F401
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()

_xls_dependency_exc_info = None
try:
    import xlrd  # noqa: F401
except ImportError:
    _xls_dependency_exc_info = sys.exc_info()

ACCEPTED_XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
]
ACCEPTED_XLSX_FILE_EXTENSIONS = [".xlsx"]

ACCEPTED_XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
ACCEPTED_XLS_FILE_EXTENSIONS = [".xls"]


def _has_border(border: Any) -> bool:
    """True if any side of the cell border has a visible style."""
    if border is None:
        return False
    return any(
        getattr(side, "style", None) not in (None, "")
        for side in (border.left, border.right, border.top, border.bottom)
    )


_IMAGE_EXT = {"jpeg": "jpg", "jpg": "jpg", "png": "png", "gif": "gif", "bmp": "bmp"}


def _image_ext(fmt: str | None) -> str:
    return _IMAGE_EXT.get((fmt or "").lower(), (fmt or "png").lower())


def _image_bytes(img: Any) -> bytes:
    """Raw bytes of an embedded image, without requiring Pillow.

    openpyxl reads workbook images with ``ref`` as a ``BytesIO`` of the raw zip
    bytes; ``getvalue()`` returns them directly (works even after openpyxl
    closes the stream). ``Image._data()`` requires Pillow and can fail on a
    closed stream, so it is only the fallback for other ref types.
    """
    ref = getattr(img, "ref", None)
    if hasattr(ref, "getvalue"):
        return ref.getvalue()
    return img._data()


def _extract_images(ws: Any, sheet_name: str) -> tuple[dict[str, bytes], dict[int, list[str]]]:
    """Extract a worksheet's embedded images.

    Returns ``(images, images_by_row)`` where ``images`` maps a relative path
    under ``assets/<sheet>/`` to raw bytes (main.py writes them there), and
    ``images_by_row`` maps a 0-based anchor row to the Markdown reference lines
    so the renderer can place each image near its anchored row.
    """
    images: dict[str, bytes] = {}
    images_by_row: dict[int, list[str]] = {}
    base = f"assets/{sanitize_filename(sheet_name)}"
    for i, img in enumerate(getattr(ws, "_images", None) or [], 1):
        try:
            row = img.anchor._from.row  # 0-based top-left anchor row
        except AttributeError:
            row = 0
        rel = f"{base}/image_{i}.{_image_ext(img.format)}"
        images[rel] = _image_bytes(img)
        images_by_row.setdefault(row, []).append(f"![{sheet_name}_{i}]({rel})")
    return images, images_by_row


def _rows_image_blocks(images_by_row: dict[int, list[str]] | None, rows) -> list[str]:
    """All image reference lines anchored within ``rows`` (row order)."""
    if not images_by_row:
        return []
    return [line for r in rows for line in images_by_row.get(r, [])]


def _table_to_markdown(header: list[str], body: list[list[str]]) -> str:
    """Build a GitHub-flavored Markdown table directly (no HTML/markdownify pass).

    The HTML→markdownify round-trip is the dominant cost for large sheets (a
    53k×27 table takes minutes), so tables are emitted directly. Pipes in cells
    are escaped; embedded newlines become ``<br>``.
    """
    if not header:
        return ""

    def cell(text: str) -> str:
        return text.replace("|", "\\|").replace("\n", "<br>")

    width = len(header)
    lines = ["| " + " | ".join(cell(h) for h in header) + " |"]
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in body:
        padded = (list(row) + [""] * width)[:width]
        lines.append("| " + " | ".join(cell(c) for c in padded) + " |")
    return "\n".join(lines)


def _render_table_region(
    cells: list[list[Any]],
    header_idx: int,
    end: int,
    lo: int,
    hi: int,
) -> str:
    """Render a table region's sub-matrix (header row + data rows) as Markdown."""
    header = [_cell_text(v) for v in cells[header_idx][lo : hi + 1]]
    body = [
        [_cell_text(v) for v in cells[r][lo : hi + 1]]
        for r in range(header_idx + 1, end + 1)
    ]
    return _table_to_markdown(header, body)


def _render_region_mixed(
    cells: list[list[Any]],
    metas,
    regions: list[tuple[int, int]],
    *,
    use_border: bool,
    images_by_row: dict[int, list[str]] | None = None,
) -> str:
    """Render one sheet mixing table regions and document prose, in row order.

    ``images_by_row`` (anchor row → Markdown image lines) is emitted at the
    corresponding row: images on prose rows appear right after that row, images
    anchored inside a table region appear after the table block.
    """
    region_of: dict[int, int] = {}
    for ordinal, (a, b) in enumerate(regions):
        for r in range(a, b + 1):
            region_of[r] = ordinal

    blocks: list[str] = []
    ridx = 0
    n = len(cells)
    while ridx < n:
        if ridx in region_of:
            a, b = regions[region_of[ridx]]
            layout = detect_region_layout(cells, metas, (a, b), use_border)
            if layout is None:  # defensive: no content/span at all → prose
                for r in range(a, b + 1):
                    items = [_cell_text(v) for v in cells[r] if _cell_text(v)]
                    if items:
                        blocks.append(render_document_row(items))
                    blocks.extend(_rows_image_blocks(images_by_row, (r,)))
                ridx = b + 1
                continue

            lo, hi, header_idx, captions = layout
            region_max = max(
                _row_content_in_span(cells[r], lo, hi) for r in range(a, b + 1)
            )
            has_data = header_idx + 1 <= b

            if region_max >= MIN_TABLE_WIDTH and has_data:
                for cap in captions:  # captions render above the table
                    caption_text = _render_caption(cells[cap], lo, hi)
                    if caption_text:
                        blocks.append(caption_text)
                blocks.append(_render_table_region(cells, header_idx, b, lo, hi))
                notes = _render_out_of_span(cells, a, b, lo, hi)
                if notes:
                    blocks.append(notes)
                blocks.extend(_rows_image_blocks(images_by_row, range(a, b + 1)))
            else:  # too thin / single-row region → prose (bordered callout, etc.)
                for r in range(a, b + 1):
                    items = [_cell_text(v) for v in cells[r] if _cell_text(v)]
                    if items:
                        blocks.append(render_document_row(items))
                    blocks.extend(_rows_image_blocks(images_by_row, (r,)))
            ridx = b + 1
        else:
            items = [_cell_text(v) for v in cells[ridx] if _cell_text(v)]
            if items:
                blocks.append(render_document_row(items))
            blocks.extend(_rows_image_blocks(images_by_row, (ridx,)))
            ridx += 1

    return "\n\n".join(x for x in blocks if x).strip()


def _sheet_to_markdown(
    name: str,
    cells: list[list[Any]],
    merged_count: int,
    kwargs: dict[str, Any],
    border_flags: list[list[bool]] | None = None,
    images_by_row: dict[int, list[str]] | None = None,
) -> str:
    """Convert one sheet's raw cells to a Markdown body (no ``## name`` heading).

    Region-based mixed rendering first: any table region (bordered / dense /
    4×4-block) renders as a Markdown table and everything else as document
    prose. When no region is found, fall back to whole-sheet classification
    (this is the path for pure documents and ``.xls`` files). Embedded images
    are emitted near their anchored row via ``images_by_row``.
    """
    metas = analyze_rows(cells, border_flags)
    dense = dense_4x4_rows(cells)
    table_flags = [is_table_row(m, dense) for m in metas]
    regions = group_table_regions(table_flags)

    # Keep only regions that can actually render as a table (wide enough and
    # with at least one data row). Thin/partial regions (e.g. a single bordered
    # label row) do not qualify — if none qualify, fall back to whole-sheet
    # classification so a sparse borderless table isn't flattened into prose.
    use_border = border_flags is not None
    qualifying = []
    for region in regions:
        layout = detect_region_layout(cells, metas, region, use_border)
        if layout is None:
            continue
        lo, hi, header_idx, _ = layout
        region_max = max(
            _row_content_in_span(cells[r], lo, hi) for r in range(region[0], region[1] + 1)
        )
        if region_max >= MIN_TABLE_WIDTH and header_idx + 1 <= region[1]:
            qualifying.append(region)

    if qualifying:
        body = _render_region_mixed(
            cells,
            metas,
            qualifying,
            use_border=use_border,
            images_by_row=images_by_row,
        )
        if kwargs.get("debug"):
            if len(qualifying) == 1 and qualifying[0] == (0, len(cells) - 1):
                print(f"[{name}] table: 1 region covering all {len(cells)} rows")
            else:
                desc = ",".join(f"{a + 1}-{b + 1}" for a, b in qualifying)
                print(f"[{name}] mixed: {len(qualifying)} table regions ({desc})")
        return body

    # No table regions → whole-sheet classification fallback.
    stats = compute_sheet_stats(
        n_rows=len(cells),
        n_cols=len(cells[0]) if cells else 0,
        merged_count=merged_count,
        sheet_name=name,
        cells=cells,
    )
    sheet_mode, sheet_reason = classify_sheet(stats)
    if kwargs.get("debug"):
        print(f"[{name}] {sheet_mode} ({sheet_reason})")

    if sheet_mode == MODE_TABLE and cells:
        header = [_cell_text(v) for v in cells[0]]
        body = [[_cell_text(v) for v in row] for row in cells[1:]]
        md = _table_to_markdown(header, body)
        img_lines = _rows_image_blocks(images_by_row, range(len(cells)))
        if img_lines:
            md += "\n\n" + "\n\n".join(img_lines)
        return md
    return render_document_sheet(cells, images_by_row=images_by_row)


def _join_sheets(sheets: dict[str, str]) -> str:
    """Concatenate ``{sheetname: body}`` with ``## sheetname`` headings."""
    return "".join(f"## {name}\n\n{body}\n\n" for name, body in sheets.items()).strip()


class XlsxConverter(DocumentConverter):
    """
    Converts XLSX files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLSX_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLSX_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:
        return DocumentConverterResult(
            markdown=_join_sheets(self.convert_sheets(file_stream, **kwargs))
        )

    def convert_sheets(
        self,
        file_stream: BinaryIO,
        **kwargs: Any,  # Options to pass to the converter
    ) -> dict[str, str]:
        """Convert each sheet to a Markdown body; returns ``{sheetname: markdown}``.

        Image references are embedded in the markdown (paths under
        ``assets/<sheet>/``); use :meth:`convert_sheets_with_assets` when the
        caller also wants the image bytes written out.
        """
        return {
            s: data["markdown"]
            for s, data in self._convert_sheets_data(file_stream, kwargs).items()
        }

    def convert_sheets_with_assets(
        self,
        file_stream: BinaryIO,
        **kwargs: Any,  # Options to pass to the converter
    ) -> dict[str, dict[str, Any]]:
        """Convert each sheet to ``{sheetname: {"markdown": str, "images": {relpath: bytes}}}``.

        ``images`` maps each relative path referenced in the markdown (e.g.
        ``assets/综述/image_1.png``) to the raw image bytes, so the caller can
        write them beside the sheet's ``.md`` file and the references resolve.
        """
        return self._convert_sheets_data(file_stream, kwargs)

    def _convert_sheets_data(
        self,
        file_stream: BinaryIO,
        kwargs: dict[str, Any],
    ) -> dict[str, dict[str, Any]]:
        """One-pass parse → ``{sheetname: {"markdown", "images"}}``.

        Reads Cell objects once (values + borders) so region detection can use
        the border signal, and embedded images once (bytes + anchor rows).
        ``data_only=True``: formula cells yield their cached value.
        """
        # Check the dependencies
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[
                1
            ].with_traceback(  # type: ignore[union-attr]
                _xlsx_dependency_exc_info[2]
            )

        file_stream.seek(0)
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        sheets: dict[str, dict[str, Any]] = {}
        for s in wb.sheetnames:
            ws = wb[s]
            cells: list[list[Any]] = []
            borders: list[list[bool]] = []
            for row in ws.iter_rows():  # Cell objects → one pass for value + border
                cells.append([cell.value for cell in row])
                borders.append([_has_border(cell.border) for cell in row])
            images, images_by_row = _extract_images(ws, s)
            body = _sheet_to_markdown(
                s,
                cells,
                len(ws.merged_cells.ranges),
                kwargs,
                border_flags=borders,
                images_by_row=images_by_row,
            )
            sheets[s] = {"markdown": body, "images": images}
        return sheets


class XlsConverter(DocumentConverter):
    """
    Converts XLS files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLS_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLS_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:
        return DocumentConverterResult(
            markdown=_join_sheets(self.convert_sheets(file_stream, **kwargs))
        )

    def convert_sheets(
        self,
        file_stream: BinaryIO,
        **kwargs: Any,  # Options to pass to the converter
    ) -> dict[str, str]:
        """Convert each sheet to a Markdown body; returns ``{sheetname: markdown}``.

        ``.xls`` has no border info in xlrd 2.x, so region detection falls back
        to content-run / dense-block signals only.
        """
        # Load the dependencies
        if _xls_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xls",
                    feature="xls",
                )
            ) from _xls_dependency_exc_info[
                1
            ].with_traceback(  # type: ignore[union-attr]
                _xls_dependency_exc_info[2]
            )

        file_stream.seek(0)
        data = file_stream.read()
        book = xlrd.open_workbook(file_contents=data)
        sheets: dict[str, str] = {}
        for s in book.sheet_names():
            sh = book.sheet_by_name(s)
            cells = [sh.row_values(r) for r in range(sh.nrows)]
            # Merged ranges need xlrd formatting_info=True (requires a real file
            # path); skipped for v1 → merged_count=0.
            sheets[s] = _sheet_to_markdown(s, cells, 0, kwargs)
        return sheets

    def convert_sheets_with_assets(
        self,
        file_stream: BinaryIO,
        **kwargs: Any,  # Options to pass to the converter
    ) -> dict[str, dict[str, Any]]:
        """Same shape as :meth:`XlsxConverter.convert_sheets_with_assets` but
        with empty ``images`` — xlrd 2.x cannot extract embedded images."""
        return {s: {"markdown": m, "images": {}} for s, m in self.convert_sheets(file_stream, **kwargs).items()}
